import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def create_fmea_excel(data, form_data):
    TESTS_LIST = [
        "DESIGN CHANGE", "DIM & WORST CASE", "SIMULATION", "REDESIGN", "CHARACTERIZE",
        "CPPP", "WORST CASING-TOLERANCE", "FUNCTIONALITY TEST", "DOE - SIGNAL LOSS TESTS",
        "OOBE & INSTALL", "SYSTEM TEST WORKFLOWS", "SPECS PERFORMANCE XPUT", "SIT E2E APP",
        "USER ABUSE - TORTURE", "HALT - HIGH ACCELERATED", "BEST-BOARDS STRIFE TEST",
        "ALT - LIFE TEST", "ROBUSTNESS", "REGS EMC", "REGS SAFETY", "USABILITY",
        "SW-FW TESTS", "DATA QUALITY", "VENDOR PQAP", "HASS MFG TESTS", "MFG LINE TESTS",
        "MFG PPA", "MAINTENANCE", "DIAGNOSABILITY", "SERVICEABILITY"
    ]

    # 1. Creem el DataFrame inicial amb les dades de la IA
    df = pd.DataFrame(data)
    
    # 2. Assegurem que les columnes de càlcul existeixin (encara que estiguin buides) 
    # per evitar el KeyError abans de guardar a Excel
    calc_cols = ['RPN', 'Priority', 'Risk Level', 'Ranking']
    for col in calc_cols:
        df[col] = ""

    # 3. Eliminem 'Sources' només si realment existeix
    if 'Sources' in df.columns:
        df = df.drop(columns=['Sources'])

    # 4. Afegim les columnes dels tests i posem les "X"
    for t in TESTS_LIST: 
        df[t] = ""
        
    for i, row in df.iterrows():
        applied = row.get('Applied_Tests', [])
        for t in applied:
            if t in df.columns:
                df.at[i, t] = "X"

    # 5. Definim l'ordre final de les columnes base (Sense 'Sources')
    base_cols = ["Part", "Failure", "Function", "Failure Mode", "Effects", "Causes", 
                 "Severity", "Occurrence", "Detectability", "Cost", "RPN", 
                 "Priority", "Risk Level", "Ranking", "Links"]
    
    file_name = f"FMEA_{form_data.get('project', 'Project')}.xlsx"
    
    # Exportem a Excel (a partir de la fila 6)
    df[base_cols + TESTS_LIST].to_excel(file_name, index=False, startrow=5)

    # 6. Obrim amb openpyxl per posar les FÓRMULES DINÀMIQUES
    wb = load_workbook(file_name)
    ws = wb.active

    # Capçalera Projecte (Files 1-4)
    info = [("PROJECT:", form_data.get("project")), ("USER NAME:", form_data.get("user_name")), 
            ("VERSION:", form_data.get("version")), ("OBJECT NAME:", form_data.get("object_name"))]
    for i, (l, v) in enumerate(info, 1):
        ws[f"A{i}"], ws[f"B{i}"] = l, v
        ws[f"A{i}"].font = Font(bold=True)

    # Format Títols (Fila 6, Blau)
    blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    for col in range(1, 16): # De la A a la O
        cell = ws.cell(row=6, column=col)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # INSERCIÓ DE FÓRMULES DINÀMIQUES
    last_row = ws.max_row
    for r in range(7, last_row + 1):
        ws[f"K{r}"] = f"=G{r}*H{r}*I{r}" # RPN
        ws[f"L{r}"] = f"=J{r}*K{r}"      # Priority
        ws[f"M{r}"] = f'=IF(K{r}<=8,"LOW",IF(K{r}<=18,"MEDIUM",IF(K{r}<=27,"URGENT","CRITICAL")))'
        ws[f"N{r}"] = f"=RANK(L{r},$L$7:$L${last_row},0)"
        
        # Link clicable
        link_cell = ws.cell(row=r, column=15)
        if link_cell.value and "http" in str(link_cell.value):
            link_cell.hyperlink = link_cell.value
            link_cell.font = Font(color="0563C1", underline="single")

    # Matriu de Tests (Columna P en endavant)
    start_test_col = 16
    colors = ["00FFFF", "00FF00", "FFFF00", "C0C0C0", "00FFFF", "FFCC99", "CC99FF"]
    
    ws.merge_cells(start_row=1, start_column=start_test_col, end_row=5, end_column=start_test_col + len(TESTS_LIST) - 1)
    top_title = ws.cell(row=1, column=start_test_col)
    top_title.value = "INVESTIGATION & TESTING"
    top_title.alignment = Alignment(horizontal="center", vertical="center")
    top_title.font = Font(bold=True, size=14)

    for i, test in enumerate(TESTS_LIST):
        col_idx = start_test_col + i
        cell = ws.cell(row=6, column=col_idx)
        cell.alignment = Alignment(textRotation=90, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 5
        c_idx = 0 if i<=8 else 1 if i<=12 else 2 if i<=17 else 3 if i<=19 else 4 if i<=23 else 5 if i<=26 else 6
        cell.fill = PatternFill(start_color=colors[c_idx], fill_type="solid")

    for r in range(7, last_row + 1):
        for c in range(start_test_col, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    wb.save(file_name)
    return file_name
