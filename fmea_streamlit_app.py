import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openai import OpenAI
import json
import datetime
import pytesseract
from PIL import Image
import pdfplumber
import pytesseract

# Optional: specify tesseract command path
# On Linux Streamlit Cloud, it is usually just "tesseract"
pytesseract.pytesseract.tesseract_cmd = "tesseract"

# -----------------------------
# OpenAI client
# -----------------------------
client = OpenAI(api_key=st.secrets["openai"]["api_key"])

st.title("AI-Assisted FMEA Generator – Powered by GPT-4.1-mini")

# -----------------------------
# Session state
# -----------------------------
for key in [
    "user_name","product_name","product_description",
    "subsystem","parts","functions","requirements","version"
]:
    if key not in st.session_state:
        st.session_state[key] = datetime.date.today() if key=="version" else ""

# -----------------------------
# Inputs
# -----------------------------
st.subheader("Project Information")

user_name = st.text_input("1. User Name", key="user_name")
product_name = st.text_input("2. Product / Prototype Name", key="product_name")
product_description = st.text_area("Product Description", key="product_description")
subsystem = st.text_input("3. Subsystem to perform FMEA", key="subsystem")

parts_input = st.text_area("4. List of Parts / Components (one per line)", key="parts")
functions_input = st.text_area("5. Functions (one per line)", key="functions")
requirements_input = st.text_area("6. Main Specs / Requirements (one per line)", key="requirements")
version = st.date_input("7. Version / Date", key="version")

st.subheader("Optional Files / Images (PDF, TXT, PNG, JPG)")
uploaded_files = st.file_uploader(
    "Upload any file with extra info for FMEA",
    type=["pdf","txt","png","jpg","jpeg"],
    accept_multiple_files=True
)

# -----------------------------
# Test columns
# -----------------------------
test_columns = [
    "INVESTIGATION & TESTING","VENDOR - PART","DESIGN CHANGE","DIM & WORST CASE",
    "SIMULATION","CHARACTERIZE","CPPP","DIAGNOSTICS","FUNCTIONALITY",
    "OOBE & INSTALL","SYSTEM TEST","SIT E2E APP","HALT","ALT",
    "ROBUSTNESS","REGS EMC","REGSSAFETY","USABILITY",
    "SW-FW TESTS","MFG TESTS","MAINTENANCE","SERVICEABILITY"
]

# -----------------------------
# Helper functions
# -----------------------------
def parse_cost(x):
    try:
        x = str(x)
        if "(" in x:
            return float(x.split("(")[1].replace(")",""))
        return float(x)
    except:
        return 1

def safe_json(text):
    try:
        start = text.find("[")
        end = text.rfind("]")
        if start != -1 and end != -1:
            return json.loads(text[start:end+1])
    except:
        pass
    return []

def extract_file_content(files):
    """Extract text from PDF, TXT, and OCR from images."""
    all_text = ""
    for f in files:
        if f.type=="application/pdf":
            with pdfplumber.open(f) as pdf:
                for page in pdf.pages:
                    all_text += page.extract_text() + "\n"
        elif f.type=="text/plain":
            all_text += f.getvalue().decode("utf-8") + "\n"
        elif f.type.startswith("image/"):
            img = Image.open(f)
            text = pytesseract.image_to_string(img)
            all_text += text + "\n"
    return all_text.strip()

# -----------------------------
# AI: Add missing essentials
# -----------------------------
def ai_add_missing(functions, requirements, parts, file_text=""):
    prompt = f"""
Product: {product_name}
Description: {product_description}

Existing functions: {functions}
Existing requirements: {requirements}
Existing parts: {parts}

Extra info from uploaded files:
{file_text}

If important functions, requirements, or parts are missing, add them.

Return JSON:

{{
"additional_functions": [],
"additional_requirements": [],
"additional_parts": []
}}
"""
    try:
        r = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role":"user","content":prompt}],
            temperature=0.2
        )
        text = r.choices[0].message.content
        data = json.loads(text[text.find("{"):text.rfind("}")+1])
        functions += data.get("additional_functions",[])
        requirements += data.get("additional_requirements",[])
        parts += data.get("additional_parts",[])
    except:
        pass
    return functions, requirements, parts

# -----------------------------
# Generate FMEA
# -----------------------------
def generate_fmea():
    functions = [f.strip() for f in functions_input.split("\n") if f.strip()]
    requirements = [r.strip() for r in requirements_input.split("\n") if r.strip()]
    parts = [p.strip() for p in parts_input.split("\n") if p.strip()]

    # Extract uploaded file content
    file_text = extract_file_content(uploaded_files) if uploaded_files else ""

    # AI adds missing essentials
    functions, requirements, parts = ai_add_missing(functions, requirements, parts, file_text)

    if not functions or not requirements:
        st.warning("Enter at least one function and requirement")
        return pd.DataFrame()

    rows = []

    for function in functions:
        prompt = f"""
Product: {product_name}
Description: {product_description}
Subsystem: {subsystem}

Parts: {parts}

Function: {function}

Requirements:
{requirements}

Extra info from uploaded files:
{file_text}

For EACH requirement generate **multiple failure scenarios** (no limit on causes).

Include:
Function, Requirement, Failure Scenario, Part, Failure Mode, End Effects, Causes, Controls, Actions, Owner, Execution Phase, Severity (1-5), Occurrence (1-4), Detectability (1-3), Estimated Cost, Tests (from list), References.

Return ONLY JSON list.
"""
        with st.spinner(f"Analyzing function: {function}"):
            response = client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role":"user","content":prompt}],
                temperature=0.3
            )

        failures = safe_json(response.choices[0].message.content)

        for f in failures:
            causes = f.get("Causes",[""])
            for cause in causes:
                S = int(f.get("Severity",3))
                O = int(f.get("Occurrence",2))
                D = int(f.get("Detectability",2))
                cost_text = f.get("Estimated Cost","Medium(1)")
                cost_val = parse_cost(cost_text)
                rpn = S*O*D

                row = {
                    "Failure Scenario":f.get("Failure Scenario",""),
                    "Function":function,
                    "Requirement":f.get("Requirement",""),
                    "Part":f.get("Part",""),
                    "Failure Mode":f.get("Failure Mode",""),
                    "End Effects of Failure":f.get("End Effects",""),
                    "Causes":cause,
                    "Current Design Controls":f.get("Controls",""),
                    "Severity (S)":S,
                    "Occurrence (O)":O,
                    "Detectability (D)":D,
                    "RPN":rpn,
                    "Priority":rpn*cost_val,
                    "Recommended Actions":",".join(f.get("Actions",[])),
                    "Owner":f.get("Owner",""),
                    "Execution Phase":f.get("Execution Phase",""),
                    "Reference Links":",".join(f.get("References",[])),
                    "Estimated Cost":cost_text
                }

                tests = f.get("tests",[])
                for t in test_columns:
                    row[t] = "X" if t in tests else ""

                rows.append(row)

    return pd.DataFrame(rows)

# -----------------------------
# Generate button
# -----------------------------
if st.button("Generate FMEA"):
    df = generate_fmea()
    if not df.empty:
        st.session_state.df = df

# -----------------------------
# Editable table
# -----------------------------
if "df" in st.session_state:
    st.subheader("Editable FMEA Table")
    edited_df = st.data_editor(st.session_state.df,use_container_width=True)

    edited_df["RPN"] = (
        edited_df["Severity (S)"] *
        edited_df["Occurrence (O)"] *
        edited_df["Detectability (D)"]
    )

    edited_df["Priority"] = edited_df["RPN"] * edited_df["Estimated Cost"].apply(parse_cost)

    st.session_state.df = edited_df

    # -----------------------------
    # Excel Export
    # -----------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "FMEA"

    headers = list(edited_df.columns)
    ws.append(headers)
    ws.row_dimensions[1].height = 60

    for i,h in enumerate(headers,1):
        c = ws.cell(1,i)
        c.font = Font(bold=True,color="FFFFFF",size=12)
        c.fill = PatternFill(start_color="4F81BD",end_color="4F81BD",fill_type="solid")
        c.alignment = Alignment(horizontal="center",vertical="center")
        ws.column_dimensions[c.column_letter].width = 25

    for r,row in enumerate(edited_df.values,2):
        for c,val in enumerate(row,1):
            if isinstance(val,list):
                val = ", ".join(val)
            if val is None:
                val = ""
            ws.cell(row=r,column=c,value=str(val))

    output = BytesIO()
    wb.save(output)

    st.download_button(
        "Download Excel",
        output.getvalue(),
        file_name=f"FMEA_{product_name}.xlsx"
    )

